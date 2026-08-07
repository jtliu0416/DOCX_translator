"""XLSX processing helpers.

Extract translatable text cells from workbooks and replace original cell values
with translated text in a copy of the original workbook.
"""

import asyncio
import json
from typing import Any

from openpyxl import load_workbook

INVALID_SHEET_TITLE_CHARS = set(":\\/?*[]")
MAX_SHEET_TITLE_LENGTH = 31


def _contains_chinese(text: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" or "\u3400" <= char <= "\u4dbf" for char in text)


def _contains_english(text: str) -> bool:
    return any(("a" <= char <= "z") or ("A" <= char <= "Z") for char in text)


def _english_ratio(text: str) -> float:
    stripped = text.strip()
    if not stripped:
        return 0.0
    english_count = sum(1 for char in stripped if ("a" <= char <= "z") or ("A" <= char <= "Z"))
    return english_count / len(stripped)


def _analyze_translatability(text: str, source_lang: str) -> dict[str, bool]:
    has_chinese = _contains_chinese(text)
    has_english = _contains_english(text)
    english_ratio = _english_ratio(text)

    if source_lang == "zh":
        has_source = has_chinese
    elif source_lang == "en":
        has_source = has_english
    else:
        has_source = bool(text.strip())

    return {
        "contains_chinese": has_chinese,
        "contains_english": has_english,
        "already_translated": False,
        "skip": not has_source,
    }


def _is_text_cell_value(value: Any) -> bool:
    return isinstance(value, str) and bool(value.strip())


def _sanitize_sheet_title(title: str, fallback: str) -> str:
    sanitized = "".join(
        "_" if char in INVALID_SHEET_TITLE_CHARS or ord(char) < 32 else char
        for char in title.strip()
    ).strip()
    if not sanitized:
        sanitized = fallback
    return sanitized[:MAX_SHEET_TITLE_LENGTH]


def _unique_sheet_title(title: str, existing_titles: set[str]) -> str:
    candidate = title
    index = 2
    while candidate.lower() in existing_titles:
        suffix = f"_{index}"
        candidate = f"{title[:MAX_SHEET_TITLE_LENGTH - len(suffix)]}{suffix}"
        index += 1
    return candidate


def _extract_cells_sync(
    xlsx_path: str,
    output_json_path: str,
    source_lang: str,
    target_lang: str,
) -> dict:
    workbook = load_workbook(xlsx_path, data_only=False)
    units = []
    index = 0

    for worksheet in workbook.worksheets:
        sheet_title = worksheet.title.strip()
        language = _analyze_translatability(sheet_title, source_lang)
        units.append({
            "index": index,
            "type": "excel_sheet",
            "sheet_name": worksheet.title,
            "text": sheet_title,
            "source_lang": source_lang,
            "target_lang": target_lang,
            **language,
        })
        index += 1

        for row in worksheet.iter_rows():
            for cell in row:
                if not _is_text_cell_value(cell.value):
                    continue

                text = cell.value.strip()
                language = _analyze_translatability(text, source_lang)
                units.append({
                    "index": index,
                    "type": "excel_cell",
                    "sheet_name": worksheet.title,
                    "coordinate": cell.coordinate,
                    "text": text,
                    "source_lang": source_lang,
                    "target_lang": target_lang,
                    **language,
                })
                index += 1

    result = {"units": units}
    with open(output_json_path, "w", encoding="utf-8") as handle:
        json.dump(result, handle, ensure_ascii=False, indent=2)
    return result


async def extract_cells(
    xlsx_path: str,
    output_json_path: str,
    source_lang: str = "zh",
    target_lang: str = "en",
) -> dict:
    return await asyncio.to_thread(
        _extract_cells_sync,
        xlsx_path,
        output_json_path,
        source_lang,
        target_lang,
    )


def _load_translations(translations_json_path: str) -> dict[int, str]:
    with open(translations_json_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    translations = {}
    for item in payload.get("translations", []):
        translations[int(item["index"])] = item.get("text", "")
    return translations


def _load_units(cells_json_path: str) -> list[dict]:
    with open(cells_json_path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return payload.get("units", [])


def _insert_translations_sync(
    xlsx_path: str,
    translations_json_path: str,
    output_path: str,
    cells_json_path: str,
) -> None:
    translations = _load_translations(translations_json_path)
    units = _load_units(cells_json_path)

    workbook = load_workbook(xlsx_path, data_only=False)
    for unit in units:
        if unit.get("type") != "excel_cell" or unit.get("skip"):
            continue

        translated_text = translations.get(int(unit["index"]))
        if not translated_text:
            continue

        worksheet = workbook[unit["sheet_name"]]
        cell = worksheet[unit["coordinate"]]
        if not _is_text_cell_value(cell.value):
            continue

        translated_text = translated_text.strip()
        if not translated_text:
            continue

        cell.value = translated_text

    existing_titles = {worksheet.title.lower() for worksheet in workbook.worksheets}
    for unit in units:
        if unit.get("type") != "excel_sheet" or unit.get("skip"):
            continue

        translated_text = translations.get(int(unit["index"]))
        if not translated_text:
            continue

        original_title = unit["sheet_name"]
        if original_title not in workbook.sheetnames:
            continue

        worksheet = workbook[original_title]
        existing_titles.discard(worksheet.title.lower())

        fallback = f"Sheet{workbook.worksheets.index(worksheet) + 1}"
        translated_title = _sanitize_sheet_title(translated_text, fallback)
        translated_title = _unique_sheet_title(translated_title, existing_titles)
        worksheet.title = translated_title
        existing_titles.add(translated_title.lower())

    workbook.save(output_path)


async def insert_cell_translations(
    xlsx_path: str,
    translations_json_path: str,
    output_path: str,
    cells_json_path: str,
) -> None:
    await asyncio.to_thread(
        _insert_translations_sync,
        xlsx_path,
        translations_json_path,
        output_path,
        cells_json_path,
    )


async def validate_xlsx(xlsx_path: str) -> bool:
    try:
        await asyncio.to_thread(load_workbook, xlsx_path, data_only=False)
        return True
    except Exception:
        return False
